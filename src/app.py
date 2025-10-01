import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from xlsx2html import xlsx2html
from playwright.sync_api import sync_playwright
import tempfile


def export_excel_sheets_to_images(xlsx_path, output_dir="./output", dpi=2):
    """
    Export each Excel sheet to high-quality PNG images.

    Args:
        xlsx_path: Path to Excel file
        output_dir: Output directory for images
        dpi: Device pixel ratio (1=normal, 2=retina, 3=ultra high-res)
    """
    xlsx_path = Path(xlsx_path)
    output_dir = Path(output_dir)

    if not xlsx_path.exists():
        print(f"Error: File not found: {xlsx_path}")
        sys.exit(1)

    output_dir.mkdir(exist_ok=True, parents=True)

    # Get all sheet names
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_names = wb.sheetnames
    print(f"\n📊 Found {len(sheet_names)} sheets: {sheet_names}\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            device_scale_factor=dpi, viewport={"width": 1920, "height": 1080}
        )
        page = context.new_page()

        success_count = 0

        for idx, sheet_name in enumerate(sheet_names, 1):
            try:
                # Create temporary HTML file
                with tempfile.NamedTemporaryFile(
                    mode="w", suffix=".html", delete=False, encoding="utf-8"
                ) as tmp:
                    tmp_path = tmp.name
                    xlsx2html(str(xlsx_path), tmp_path, sheet=sheet_name)

                # Read and style the HTML
                with open(tmp_path, "r", encoding="utf-8") as f:
                    html_content = f.read()

                styled_html = f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="utf-8">
                    <style>
                        body {{
                            margin: 20px;
                            background: white;
                            font-family: Arial, sans-serif;
                        }}
                        table {{
                            border-collapse: collapse;
                            background: white;
                        }}
                        td, th {{
                            border: 1px solid #ccc;
                            padding: 5px 8px;
                            font-size: 12px;
                            white-space: nowrap;
                        }}
                    </style>
                </head>
                <body>
                    {html_content}
                </body>
                </html>
                """

                with open(tmp_path, "w", encoding="utf-8") as f:
                    f.write(styled_html)

                # Load and screenshot
                page.goto(f"file://{tmp_path}")
                page.wait_for_load_state("networkidle")

                table = page.locator("table").first
                if table:
                    safe_name = sheet_name.replace(" ", "_").replace("/", "_")
                    screenshot_path = output_dir / f"{safe_name}.png"
                    table.screenshot(path=str(screenshot_path))
                    print(
                        f"[{idx}/{len(sheet_names)}] {sheet_name} -> {screenshot_path.name}"
                    )
                    success_count += 1
                else:
                    print(f"[{idx}/{len(sheet_names)}] {sheet_name} -> No table found")

                # Cleanup
                os.unlink(tmp_path)

            except Exception as e:
                print(f"[{idx}/{len(sheet_names)}] {sheet_name} -> Error: {e}")
                continue

        browser.close()

        print(
            f"\n✅ Complete: {success_count}/{len(sheet_names)} sheets exported to {output_dir.absolute()}\n"
        )


if __name__ == "__main__":
    # Configuration
    EXCEL_FILE = "./src/Test_Report_Test_Report_8206853 Test report mill_8206853.xlsx"
    OUTPUT_DIR = "./output"
    DPI = 2  # 1=normal, 2=high-res, 3=ultra high-res

    export_excel_sheets_to_images(EXCEL_FILE, OUTPUT_DIR, DPI)
