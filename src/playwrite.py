import os
from pathlib import Path
from openpyxl import load_workbook
from xlsx2html import xlsx2html
from playwright.sync_api import sync_playwright
import tempfile
import shutil


def export_excel_sheets_to_images(xlsx_path, output_dir="./output", dpi=2):
    """
    Export each sheet from Excel to high-quality PNG images.

    Args:
        xlsx_path: Path to Excel file
        output_dir: Output directory for images
        dpi: Device pixel ratio (2 = retina/high-res, 1 = normal)
    """
    xlsx_path = Path(xlsx_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True, parents=True)

    # Get all sheet names
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Found {len(sheet_names)} sheets: {sheet_names}")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(device_scale_factor=dpi)
        page = context.new_page()

        for sheet_name in sheet_names:
            try:
                # Create temporary HTML file for this sheet
                with tempfile.NamedTemporaryFile(
                    mode="w", suffix=".html", delete=False, encoding="utf-8"
                ) as tmp:
                    tmp_path = tmp.name

                    # Convert Excel sheet to HTML
                    xlsx2html(str(xlsx_path), tmp_path, sheet=sheet_name)

                # Read the HTML and add styling for better screenshots
                with open(tmp_path, "r", encoding="utf-8") as f:
                    html_content = f.read()

                # Add CSS for better rendering
                styled_html = f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="utf-8">
                    <style>
                        body {{
                            margin: 20px;
                            background: white;
                            font-family: Arial, sans-serif;
                        }}
                        table {{
                            border-collapse: collapse;
                            background: white;
                        }}
                        td, th {{
                            border: 1px solid #ccc;
                            padding: 5px 8px;
                            font-size: 12px;
                        }}
                    </style>
                </head>
                <body>
                    {html_content}
                </body>
                </html>
                """

                # Write styled HTML back
                with open(tmp_path, "w", encoding="utf-8") as f:
                    f.write(styled_html)

                # Load HTML in browser
                page.goto(f"file://{tmp_path}")
                page.wait_for_load_state("networkidle")

                # Get the table element and screenshot only that
                table = page.locator("table").first
                if table:
                    safe_name = sheet_name.replace(" ", "_").replace("/", "_")
                    screenshot_path = output_dir / f"{safe_name}.png"
                    table.screenshot(path=str(screenshot_path))
                    print(f"Saved: {screenshot_path}")
                else:
                    print(f"No table found in sheet: {sheet_name}")

                # Clean up temp file
                os.unlink(tmp_path)

            except Exception as e:
                print(f"Error processing sheet {sheet_name}: {e}")
                continue

        browser.close()


# Usage
if __name__ == "__main__":
    export_excel_sheets_to_images(
        "./src/Test_Report_Test_Report_8206853 Test report mill_8206853.xlsx",
        "./output",
        dpi=2,  # Use 2 for high-res, 3 for ultra high-res
    )
